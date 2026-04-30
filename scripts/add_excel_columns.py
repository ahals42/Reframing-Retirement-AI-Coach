"""Add 'What Was Done' and 'New Expected Behaviour' columns to reviewer_notes_by_theme.xlsx."""

import openpyxl
from openpyxl.styles import Alignment, Font

PATH = "docs/reviewer_notes_by_theme.xlsx"

# Keyed by (sheet_name, data_row) where data_row=1 = first non-header row
# Each entry: (what_was_done, new_expected_behaviour)

DATA = {

# ─── T1 INTERVENTION GROUNDING ────────────────────────────────────────────────

("T1 Intervention", 1): (
    "Added Lesson 1 to topic-to-lesson mapping in BASE_PROMPT; set allow_module_references=True in default mode.",
    "Response now ends with a natural reference to Lesson 1 when discussing activity guidelines or why PA matters.",
),
("T1 Intervention", 2): (
    "Updated BASE_PROMPT lesson reference guide: instruct model to state fact first, then reference lesson — not lead with 'did you know?'",
    "Response states the relevant fact (e.g., 'Exercise supports memory') then adds 'You can explore this in Lesson 2.'",
),
("T1 Intervention", 3): (
    "Added Lesson 2 to topic-to-lesson mapping in BASE_PROMPT for cognitive benefits.",
    "Response on concentration/memory now ends with a reference to Lesson 2.",
),
("T1 Intervention", 4): (
    "Added Lesson 2 to topic-to-lesson mapping for mood/affective attitude topics.",
    "Response on mood benefits now ends with a reference to Lesson 2.",
),
("T1 Intervention", 5): (
    "Same as row 4 — Lesson 2 mapped for mood/affective topics.",
    "Same as row 4.",
),
("T1 Intervention", 6): (
    "Added Lessons 1-3 to topic-to-lesson mapping for perceived capability topics.",
    "Response on capability/realism of being active now references Lessons 1-3.",
),
("T1 Intervention", 7): (
    "Added Lesson 3 to topic-to-lesson mapping for confidence/capability topics.",
    "Response on confidence now ends with a reference to Lesson 3.",
),
("T1 Intervention", 8): (
    "Changed 'Slide' to 'Page' in response mode instructions in agent.py and module reference instruction.",
    "All lesson citations now say 'page' not 'slide' (e.g., 'Lesson 6, page 8').",
),
("T1 Intervention", 9): (
    "Added Lesson 4 to topic-to-lesson mapping for planning/goal-setting topics.",
    "Response on activity planning now references Lesson 4.",
),
("T1 Intervention", 10): (
    "Added Lesson 5 to topic-to-lesson mapping for self-monitoring topics.",
    "Response on tracking activity now references Lesson 5.",
),
("T1 Intervention", 11): (
    "Added Lesson 5 to topic-to-lesson mapping for social monitoring topics.",
    "Response on tracking with a partner now references Lesson 5.",
),
("T1 Intervention", 12): (
    "Same as row 11.",
    "Same as row 11.",
),
("T1 Intervention", 13): (
    "Added Lessons 5 and 6 to topic-to-lesson mapping for distraction/focus topics.",
    "Response on staying focused now references Lessons 5 and 6.",
),
("T1 Intervention", 14): (
    "Added Lesson 6 to topic-to-lesson mapping for reactive regulation topics.",
    "Response on 'don't feel like it' now references Lesson 6.",
),
("T1 Intervention", 15): (
    "Updated BASE_PROMPT: instruct model to give strategies first then reference lesson (not lead with lesson).",
    "Response gives concrete emotion regulation strategies, then ends with reference to Lesson 6.",
),
("T1 Intervention", 16): (
    "Added Lesson 6 to topic-to-lesson mapping for self-talk/motivational dialogue.",
    "Response on positive self-talk now references Lesson 6.",
),
("T1 Intervention", 17): (
    "Changed 'Slide' to 'Page' in retriever.py reference() and agent.py instructions. Phrasing aligned with Lesson 6.",
    "Citation reads 'Lesson 6, page 8' not 'Lesson 6, Slide 8'; self-talk phrasing matches lesson.",
),
("T1 Intervention", 18): (
    "Changed 'Slide' to 'Page' in retriever.py reference() for lesson chunks.",
    "Citation reads 'Lesson 7, page X' not 'Lesson 7, Slide X'.",
),
("T1 Intervention", 19): (
    "Fixed science module name in retriever.py: 'Science Module X' → 'The Science Behind the Lessons'. Added Lesson 7 to lesson mapping.",
    "Citation reads 'The Science Behind the Lessons' and Lesson 7 is also referenced for habit content.",
),
("T1 Intervention", 20): (
    "Fixed science module name in retriever.py; corrected page reference for Lesson 7 cue content to pages 9-11.",
    "Citation reads 'Lesson 7, pages 9-11: [title]' with correct page numbers.",
),
("T1 Intervention", 21): (
    "Changed 'Slide' to 'Page' in retriever.py reference() for lesson chunks.",
    "All lesson citations say 'page' not 'slide'.",
),
("T1 Intervention", 22): (
    "Added Lessons 7 and 8 to topic-to-lesson mapping. Added qualification for habit timeline ('one study found...').",
    "Response on habit formation references Lessons 7/8; timeline stated as 'research suggests' not as a universal rule.",
),
("T1 Intervention", 23): (
    "Updated lesson mapping to cite both Lessons 7 and 8 for habit formation/timeline content.",
    "Response on habit timelines references both Lessons 7 and 8.",
),
("T1 Intervention", 24): (
    "Added Lesson 8 to topic-to-lesson mapping for habit disruption/recovery topics.",
    "Response on breaking/recovering habits now references Lesson 8.",
),
("T1 Intervention", 25): (
    "Added Lesson 8 to topic-to-lesson mapping for habit longevity/maintenance topics.",
    "Response on making habits last now references Lesson 8.",
),
("T1 Intervention", 26): (
    "Added Lesson 9 to topic-to-lesson mapping for identity topics. Updated follow-up question to be identity-specific.",
    "Response references Lesson 9; follow-up asks specifically about seeing oneself as an active person.",
),
("T1 Intervention", 27): (
    "Fixed lesson mapping: values content now maps to Lesson 10, not the previously incorrect lesson.",
    "Citation reads 'Lesson 10' for values/identity content.",
),
("T1 Intervention", 28): (
    "Added Lesson 10 to topic-to-lesson mapping for values and activity topics.",
    "Response on values and exercise now references Lesson 10.",
),
("T1 Intervention", 29): (
    "Added instruction in BASE_PROMPT to use lesson-consistent activity examples (walking, stretching, gardening for belonging context); added healthcare referral for physical limitation queries.",
    "Examples align with lesson content; knee-related queries also suggest consulting healthcare provider.",
),
("T1 Intervention", 30): (
    "Added Lesson 1 to topic-to-lesson mapping for strength training safety topics.",
    "Response on weight training safety references Lesson 1 and suggests consulting a fitness professional.",
),
("T1 Intervention", 31): (
    "Added Lesson 4 to topic-to-lesson mapping for goal reminder/memory topics.",
    "Response on not forgetting goals now references Lesson 4.",
),
("T1 Intervention", 32): (
    "Added Lesson 5 to topic-to-lesson mapping for staying on track/distraction topics.",
    "Response on staying on track now references Lesson 5.",
),
("T1 Intervention", 33): (
    "Fixed lesson mapping for emotion regulation: now correctly maps to Lesson 6, removing incorrect Lesson 10 reference.",
    "Emotion regulation citation reads 'Lesson 6' not 'Lesson 10, Slide 23'.",
),
("T1 Intervention", 34): (
    "Changed 'Slide' to 'Page' in retriever.py and agent.py instructions.",
    "All citations use 'page' terminology.",
),
("T1 Intervention", 35): (
    "Fixed science module name in retriever.py. Added Lesson 7 as co-reference for habit content.",
    "Citations correctly read 'The Science Behind the Lessons' and include Lesson 7.",
),
("T1 Intervention", 36): (
    "Added Lesson 9 to topic-to-lesson mapping; updated follow-up to identity-specific phrasing.",
    "Response references Lesson 9; follow-up focuses on identity not just positive experience.",
),
("T1 Intervention", 37): (
    "Added 'M-PAC framework' enforcement to BASE_PROMPT; fixed science module naming in retriever.py.",
    "Responses say 'M-PAC framework' not 'the M-PAC'; science lesson named correctly.",
),
("T1 Intervention", 38): (
    "Fixed science module name in retriever.py; added M-PAC → science module mapping in lesson guide.",
    "M-PAC explanation now references 'The Science Behind the Lessons'.",
),
("T1 Intervention", 39): (
    "Fixed science module name in retriever.py; added reflective constructs to science module mapping.",
    "Instrumental beliefs citation reads 'The Science Behind the Lessons'.",
),
("T1 Intervention", 40): (
    "Fixed science module name in retriever.py; changed 'Slide' to 'Page'.",
    "Citation reads 'The Science Behind the Lessons, page X: [title]'.",
),
("T1 Intervention", 41): (
    "Fixed science module name in retriever.py; added both science module and Lesson 3 to perceived capability mapping.",
    "Perceived capability responses cite both 'The Science Behind the Lessons' and Lesson 3.",
),
("T1 Intervention", 42): (
    "Fixed science module name and corrected page reference to page 27 in retriever.py / lesson mapping.",
    "Affective judgements citation reads 'The Science Behind the Lessons, page 27'.",
),
("T1 Intervention", 43): (
    "Fixed science module name in retriever.py; mapped regulatory phase to second science module.",
    "Regulatory phase citation reads 'The Science Behind the Lessons' (second module).",
),
("T1 Intervention", 44): (
    "Fixed lesson reference for regulatory phase to second science module in lesson mapping.",
    "Same as row 43.",
),
("T1 Intervention", 45): (
    "Fixed science module name; corrected regulatory phase description to 'translating intentions into action' in BASE_PROMPT.",
    "Regulatory phase citation correct; description covers planning, self-monitoring, emotion regulation.",
),
("T1 Intervention", 46): (
    "Added Lesson 4 and second science module to topic-to-lesson mapping for intention-behaviour gap.",
    "Intention-behaviour gap responses reference Lesson 4.",
),
("T1 Intervention", 47): (
    "Fixed science module name in retriever.py; changed 'Slide' to 'Page'.",
    "Social monitoring science citation reads 'The Science Behind the Lessons, page X'.",
),
("T1 Intervention", 48): (
    "Added Lessons 7 and 8 to topic-to-lesson mapping for habit persistence.",
    "Habit-sticking responses reference Lessons 7 and 8.",
),
("T1 Intervention", 49): (
    "Fixed science module name in retriever.py for habit measurement content.",
    "Self-reported habit index citation reads 'The Science Behind the Lessons'.",
),
("T1 Intervention", 50): (
    "Added qualification in BASE_PROMPT: asymptotic curve varies by person; do not overstate 'starts rapidly'.",
    "Asymptotic curve described with qualifier: 'improvement tends to be faster early on, though this varies by person'.",
),
("T1 Intervention", 51): (
    "Fixed science module name in retriever.py for affect content.",
    "Affect definition citation reads 'The Science Behind the Lessons'.",
),
("T1 Intervention", 52): (
    "Fixed science module name; added Lesson 6 as secondary reference for affect.",
    "Affect response cites science module and Lesson 6.",
),
("T1 Intervention", 53): (
    "Fixed science module name and corrected page reference to page 12 for valence.",
    "Valence citation reads 'The Science Behind the Lessons, page 12'.",
),
("T1 Intervention", 54): (
    "Updated hedonic motivation description in BASE_PROMPT to match lesson wording ('drive for pleasure and enjoyment'); changed 'Slide' to 'Page'.",
    "Hedonic motivation described as 'drive for pleasure and enjoyment'; citation uses 'page'.",
),
("T1 Intervention", 55): (
    "Corrected page reference for identity/exercise content to pages 5 and 6 in lesson mapping.",
    "Identity research citation reads 'pages 5 and 6' correctly.",
),
("T1 Intervention", 56): (
    "Fixed science module name; added Lesson 10 as secondary reference for ACT.",
    "ACT evidence cites 'The Science Behind the Lessons' and Lesson 10.",
),
("T1 Intervention", 57): (
    "Added note in BASE_PROMPT that M-PAC is covered across all three science modules.",
    "M-PAC explanation notes it is covered across all 'The Science Behind the Lessons' modules.",
),
("T1 Intervention", 58): (
    "Fixed science module name to first science module for perceived capability.",
    "Perceived capability detail cites 'The Science Behind the Lessons, Lesson 1'.",
),
("T1 Intervention", 59): (
    "Fixed science module name to second science module for regulatory phase.",
    "Regulatory phase detail cites 'The Science Behind the Lessons, Lesson 2'.",
),
("T1 Intervention", 60): (
    "Fixed science module name for intention-behaviour gap.",
    "Intention-behaviour gap cites 'The Science Behind the Lessons'.",
),
("T1 Intervention", 61): (
    "Fixed science module name; changed 'Slide' to 'Page'; added Lesson 6 as secondary reference for affect.",
    "Affect definition cites 'The Science Behind the Lessons, page X' and Lesson 6.",
),
("T1 Intervention", 62): (
    "Fixed science module name and page reference to page 24 for ACT research.",
    "ACT research citation reads 'The Science Behind the Lessons, page 24'.",
),
("T1 Intervention", 63): (
    "Added Lesson 4 to topic-to-lesson mapping for goal-setting follow-up queries.",
    "Goal-setting response now references Lesson 4.",
),

# ─── T2 RESOURCE SCOPE ────────────────────────────────────────────────────────

("T2 Resource Scope", 1): (
    "Removed 'mention at least one concrete option by name' from default retrieval instruction in agent.py; added instruction to direct to Resources section.",
    "Response directs user to 'What is going on in your area?' in Resources section instead of naming specific pools or programs.",
),
("T2 Resource Scope", 2): (
    "Same retrieval instruction change; added gender-neutral resource description instruction (T5 Change 5A).",
    "At-home resource response leads with Resources section direction, not a specific video; description is gender-neutral.",
),
("T2 Resource Scope", 3): (
    "Retrieval instruction updated to prohibit naming specific programs; direction to Resources section added.",
    "Response says 'browse local options in Resources > What is going on in your area?' instead of naming Silver Threads.",
),
("T2 Resource Scope", 4): (
    "Added schedule disclaimer to all activity response templates: 'Please check the Resources section or contact the centre to confirm current schedules'.",
    "Any specific schedule details are followed by a disclaimer to verify before attending.",
),
("T2 Resource Scope", 5): (
    "Same as row 3 — retrieval instruction prohibits naming specific programs.",
    "Same as row 3.",
),
("T2 Resource Scope", 6): (
    "Same as row 4 — schedule disclaimer added.",
    "Same as row 4.",
),
("T2 Resource Scope", 7): (
    "Added location clarification step to BASE_PROMPT: ask user's preferred area before suggesting local activities.",
    "Response asks 'Which neighbourhood or area of Victoria works best for you?' before directing to Resources section.",
),
("T2 Resource Scope", 8): (
    "Added explicit prohibition on external websites in BASE_PROMPT hard boundaries; retrieval instruction updated.",
    "Response directs to in-app Resources section; never mentions Meetup, external websites, or third-party platforms.",
),
("T2 Resource Scope", 9): (
    "Same as row 8 — retrieval instruction prohibits naming specific programs or external platforms.",
    "Response directs to Resources section instead of naming Oaklands Park or specific schedules.",
),
("T2 Resource Scope", 10): (
    "Same as row 4 — schedule disclaimer added.",
    "Same as row 4.",
),
("T2 Resource Scope", 11): (
    "Same as row 4 — schedule disclaimer added.",
    "Same as row 4.",
),
("T2 Resource Scope", 12): (
    "Same as row 3 — retrieval instruction prohibits naming specific programs.",
    "Same as row 3.",
),
("T2 Resource Scope", 13): (
    "Same as row 4 — schedule disclaimer added.",
    "Same as row 4.",
),
("T2 Resource Scope", 14): (
    "Same as row 4 — schedule disclaimer added.",
    "Same as row 4.",
),
("T2 Resource Scope", 15): (
    "Retrieval instruction updated to prohibit naming specific programs; Lesson 1 added to mapping for strength training.",
    "Response directs to Resources section; also references Lesson 1 for strength training content.",
),
("T2 Resource Scope", 16): (
    "Same as row 4 — schedule disclaimer added.",
    "Same as row 4.",
),
("T2 Resource Scope", 17): (
    "Same as row 4 — schedule disclaimer added.",
    "Same as row 4.",
),
("T2 Resource Scope", 18): (
    "Same as row 3 — retrieval instruction prohibits naming specific programs.",
    "Response directs to Saanich section of Resources instead of naming 'Free Easy Walks'.",
),
("T2 Resource Scope", 19): (
    "Same as row 8 — external website prohibition; direction to Resources section.",
    "Response directs to Resources section and adds schedule disclaimer.",
),
("T2 Resource Scope", 20): (
    "Same as row 4 — schedule disclaimer added; specific day removed from response.",
    "Specific day (Monday) not stated; user directed to Resources section for current schedule.",
),
("T2 Resource Scope", 21): (
    "Same as row 4 — schedule disclaimer added.",
    "Same as row 4.",
),
("T2 Resource Scope", 22): (
    "Same as row 3 — retrieval instruction prohibits naming specific programs.",
    "Response directs to Resources section instead of naming specific class venues.",
),
("T2 Resource Scope", 23): (
    "Same as row 4 — schedule disclaimer added; specific days removed.",
    "Specific days not stated; user directed to Resources section for current schedule.",
),
("T2 Resource Scope", 24): (
    "Updated home_resources instruction to use app-visible label names only ('Individual Video', 'Video Playlist', 'Blog').",
    "At-home resources use correct section labels that match what users see in the app.",
),
("T2 Resource Scope", 25): (
    "Fixed at-home resource category labels in home_resources instruction; description clarified as 'video links and articles'.",
    "Exercise categories correctly labeled; resources described as videos and articles not written content.",
),
("T2 Resource Scope", 26): (
    "Home_resources instruction updated to always establish location: 'In Resources > What Can You Do At Home'.",
    "All at-home resource mentions are prefaced with their location in the app.",
),
("T2 Resource Scope", 27): (
    "Fixed blog numbering in home data; added gender-neutral description instruction (T5 Change 5A).",
    "Blog reference correctly numbered; description is gender-neutral ('beginner flexibility routine' not 'for women over 50').",
),
("T2 Resource Scope", 28): (
    "Same as row 26 — home_resources instruction establishes Resources section location.",
    "Same as row 26.",
),
("T2 Resource Scope", 29): (
    "Updated home_resources instruction to direct to section rather than prescribing specific videos.",
    "Response directs to 'What Can You Do At Home?' section for user to browse, rather than naming specific videos.",
),
("T2 Resource Scope", 30): (
    "No code change for trip planning. Out-of-scope redirect updated (T3 Change 3B) to be a warm decline.",
    "Trip planning query gets a brief warm decline; no forced PA redirect.",
),

# ─── T3 SAFETY ────────────────────────────────────────────────────────────────

("T3 Safety", 1): (
    "Added instruction in BASE_PROMPT: when user explicitly mentions depression or anxiety, include mental health professional referral alongside PA discussion.",
    "Depression mention triggers: PA content + 'Talking with your doctor or a counsellor can be helpful for this too'.",
),
("T3 Safety", 2): (
    "Added instruction in BASE_PROMPT: avoid clinical MH terms (depression, anxiety) unless user uses them first; use 'low mood', 'feeling flat' instead.",
    "Mood-related responses use softer language; clinical terms not introduced by the chatbot unprompted.",
),
("T3 Safety", 3): (
    "Same as row 1 — depression query triggers MH referral.",
    "Same as row 1.",
),
("T3 Safety", 4): (
    "Added instruction in BASE_PROMPT: for new exercise programs (especially strength training), recommend consulting a fitness professional.",
    "Weight training safety response includes 'consulting a fitness professional or instructor is a good first step'.",
),
("T3 Safety", 5): (
    "Added MEDICAL QUERIES instruction to BASE_PROMPT: medication queries must include 'your doctor is the best resource for this'.",
    "Antidepressant query: 'Questions about your medication are best discussed directly with your doctor. If you're interested in how PA can support wellbeing, I'm happy to help with that too.'",
),
("T3 Safety", 6): (
    "Same as row 5 — medication hard boundary added to BASE_PROMPT.",
    "Same as row 5.",
),
("T3 Safety", 7): (
    "Same as row 5.",
    "Same as row 5.",
),
("T3 Safety", 8): (
    "Same as row 5.",
    "Same as row 5.",
),
("T3 Safety", 9): (
    "Same as row 5 — medication change queries refer to healthcare provider.",
    "Medication change query: healthcare provider referral before any PA discussion.",
),
("T3 Safety", 10): (
    "Same as row 9.",
    "Same as row 9.",
),
("T3 Safety", 11): (
    "Same as row 9.",
    "Same as row 9.",
),
("T3 Safety", 12): (
    "Updated tone guidance in BASE_PROMPT: acknowledge PAR-Q pre-clearance; soften cautionary language for condition-specific questions.",
    "Heart condition response: 'Since everyone's situation is different, your doctor is the best resource. If you've been cleared for activity, I can help find something that works for you.'",
),
("T3 Safety", 13): (
    "Added cardiovascular conditions to MEDICAL QUERIES list in BASE_PROMPT.",
    "Same as row 12.",
),
("T3 Safety", 14): (
    "Same as rows 12-13.",
    "Same as rows 12-13.",
),
("T3 Safety", 15): (
    "Same as rows 12-13.",
    "Same as rows 12-13.",
),
("T3 Safety", 16): (
    "Updated BASE_PROMPT: for cancer queries, do not prescribe exercise intensity; acknowledge individual variation; refer to doctor.",
    "Cancer query: 'Since everyone's situation varies by type and stage, your doctor is the best resource for what level of activity is right for you.'",
),
("T3 Safety", 17): (
    "Same as row 16 — serious health condition pattern added.",
    "Same as row 16.",
),
("T3 Safety", 18): (
    "Added diabetes/chronic disease management to MEDICAL QUERIES list; 'cure' queries trigger healthcare referral.",
    "Diabetes 'cure' query: brief note that PA supports overall health + 'for managing your condition, your doctor is the best resource'.",
),
("T3 Safety", 19): (
    "Updated out-of-scope redirect in BASE_PROMPT: warm decline with no forced PA redirect for diet/weight queries.",
    "Diet query: 'I'm not able to help with diet advice, but I hope you find something that works. If you'd like support with physical activity, I'm happy to help.'",
),
("T3 Safety", 20): (
    "Same update as row 19 — out-of-scope decline is warm and does not force a PA pivot.",
    "Calorie query: warm decline, then optional offer of PA help — not a forced redirect.",
),
("T3 Safety", 21): (
    "Added persistent health symptoms to MEDICAL QUERIES list in BASE_PROMPT.",
    "Fatigue query: acknowledge, recommend consulting doctor, then briefly note PA can support energy.",
),
("T3 Safety", 22): (
    "Same as row 21.",
    "Same as row 21.",
),
("T3 Safety", 23): (
    "Added supplements/vitamins to MEDICAL QUERIES list in BASE_PROMPT.",
    "Supplement query: 'Your doctor or pharmacist is the best resource for supplement advice.'",
),
("T3 Safety", 24): (
    "Same as row 23.",
    "Same as row 23.",
),
("T3 Safety", 25): (
    "Added trip/travel to explicit out-of-scope list; out-of-scope redirect updated to warm decline with no PA redirect.",
    "Trip planning: 'I'm not able to help with trip planning, but I hope you have a wonderful trip! If you'd like support with physical activity, I'm happy to help.'",
),
("T3 Safety", 26): (
    "Updated out-of-scope redirect: single brief decline, no repeated PA redirects, no follow-up questions for clearly unrelated queries.",
    "Restaurant query: 'I'm not able to help with that, but I hope you enjoy your meal!' — no follow-up questions.",
),

# ─── T4 CONSTRUCT PRECISION ───────────────────────────────────────────────────

("T4 Construct", 1): (
    "Added concrete Lesson 6 strategies to BASE_PROMPT for emotion regulation; fixed lesson reference to Lesson 6.",
    "Emotion management response lists specific strategies (scheduled activity, self-talk, reframing) and references Lesson 6.",
),
("T4 Construct", 2): (
    "Added construct boundary rule to BASE_PROMPT: breathing exercises in this context = calming technique only, not physiological performance.",
    "Breathing exercise response stays within emotional coping frame; no BPM counts or VO2 mentions.",
),
("T4 Construct", 3): (
    "Added cue-response link as mandatory element in habit explanations in BASE_PROMPT.",
    "Habit explanation includes: 'A habit forms when a consistent cue reliably triggers the same behaviour, making it more automatic over time.'",
),
("T4 Construct", 4): (
    "Replaced 'mental pathways' with 'cue-behaviour link' in BASE_PROMPT construct terminology.",
    "Habit mechanism described as 'cue-behaviour link' or 'cue-response connection', not 'mental pathways'.",
),
("T4 Construct", 5): (
    "Updated BASE_PROMPT M-PAC explanation to include all three phases: reflective, regulatory, reflexive.",
    "M-PAC overview covers all three phases when user asks.",
),
("T4 Construct", 6): (
    "Updated regulatory phase description in BASE_PROMPT: 'translating intentions into action through planning, self-monitoring, and emotion regulation'.",
    "Regulatory phase no longer described as just 'overcoming barriers'; full role is explained.",
),
("T4 Construct", 7): (
    "Added construct boundary: fatigue is physiological, not affective — do not conflate in emotion regulation context.",
    "Emotion regulation discussion distinguishes emotional states from physical fatigue.",
),
("T4 Construct", 8): (
    "Added automaticity as the defining marker of habit formation in BASE_PROMPT.",
    "Habit tracking described as: 'A habit is forming when you find yourself starting in response to the cue without deciding to.'",
),
("T4 Construct", 9): (
    "Fixed lesson mapping: regulatory phase now correctly attributed to second science module (lessons 4-6).",
    "Regulatory phase content referenced to 'The Science Behind the Lessons' (second module), not first.",
),

# ─── T5 PERSONALISATION ───────────────────────────────────────────────────────

("T5 Personalisation", 1): (
    "Updated BASE_PROMPT capabilities framing; added affirming older adult framing per Lesson 3.",
    "Strength training response leads with 'Older adults can absolutely strength train' rather than conservative caution.",
),
("T5 Personalisation", 2): (
    "Added gender-neutral resource description instruction to home_resources mode in agent.py.",
    "At-home resource described as 'a beginner flexibility routine for older adults', not 'for women over 50'.",
),
("T5 Personalisation", 3): (
    "Added instruction in BASE_PROMPT: do not echo user's age back as framing; affirm without anchoring on age.",
    "Response to '65 too old for stretching?' reads 'It's never too late to start' without repeating '65'.",
),
("T5 Personalisation", 4): (
    "Added activity example diversification instruction to BASE_PROMPT; pickleball restricted to user-initiated mentions.",
    "Activity examples use walking, cycling, swimming, yoga etc. instead of defaulting to pickleball.",
),
("T5 Personalisation", 5): (
    "Same as row 4.",
    "Same as row 4.",
),
("T5 Personalisation", 6): (
    "Added instruction in BASE_PROMPT: cite ACT research generically, not with university student example.",
    "ACT evidence referenced as 'research on health behaviour change' without the university student detail.",
),
("T5 Personalisation", 7): (
    "Same as row 6.",
    "Same as row 6.",
),
("T5 Personalisation", 8): (
    "Same as row 6.",
    "Same as row 6.",
),
("T5 Personalisation", 9): (
    "Added instruction in BASE_PROMPT: ask about skill level before assuming beginner status.",
    "Pickleball response asks 'Are you familiar with the game, or would you be starting fresh?' before describing beginner-friendly aspects.",
),
("T5 Personalisation", 10): (
    "Same as rows 2/3 — gender-neutral at-home resource descriptions.",
    "Same as row 2.",
),

# ─── T6 SOURCE TRACEABILITY ───────────────────────────────────────────────────

("T6 Source", 1): (
    "Added prohibition on specific numerical guidance not in lessons (BPM etc.) to BASE_PROMPT hard boundaries.",
    "Music/walking pace query answered without BPM figures; general guidance only ('music you enjoy').",
),
("T6 Source", 2): (
    "Same as row 1 for BPM. Canadian spelling enforcement added to BASE_PROMPT and _replace_em_dash() in agent.py.",
    "No BPM guidance; 'favourite' spelled correctly in all responses.",
),
("T6 Source", 3): (
    "Added qualification language instruction to BASE_PROMPT: use 'some people find...' not definitive 'easiest'.",
    "Tracking method response: 'Some people find a wall calendar the simplest way to start — others prefer a notebook or app. It depends on what you'll actually use.'",
),
("T6 Source", 4): (
    "Added sleep to explicit out-of-scope list in BASE_PROMPT; no general sleep hygiene advice permitted.",
    "Sleep query: 'Sleep is outside what I'm able to help with — I hope you find something that works. If you're interested in how PA supports overall wellbeing, I'd be glad to help.'",
),
("T6 Source", 5): (
    "Same as row 4.",
    "Same as row 4.",
),
("T6 Source", 6): (
    "Added equipment/shoe purchasing to explicit out-of-scope list in BASE_PROMPT.",
    "Shoe query: 'Finding the right shoes is outside what I can help with, but I hope you find a great pair! If you'd like help getting started with some movement, I'm happy to help.'",
),
("T6 Source", 7): (
    "Same as row 6. Out-of-scope update also ensures no follow-up questions after unrelated query decline.",
    "Same as row 6; single decline with no follow-up.",
),
("T6 Source", 8): (
    "Same as rows 6-7 — shoe/equipment queries added to hard out-of-scope enforcement.",
    "Same as row 6.",
),

# ─── T7 STYLE ─────────────────────────────────────────────────────────────────

("T7 Style", 1): (
    "Added instruction in BASE_PROMPT: when citing 150 min/week, include 1-2 examples of moderate intensity PA.",
    "150 min/week response: '...like brisk walking or easy cycling.' Examples always included.",
),
("T7 Style", 2): (
    "Added weather qualifier instruction in BASE_PROMPT for outdoor activity discussions.",
    "Outdoor walking comparison: 'Walking outside, when the weather cooperates, tends to provide more mood benefits...'",
),
("T7 Style", 3): (
    "No change — reviewer approved this response.",
    "Response continues to perform as-is.",
),
("T7 Style", 4): (
    "Added instruction in BASE_PROMPT: do not re-define or re-expand an acronym already introduced in this session.",
    "ACT not re-defined in follow-up turn if already introduced; used directly without expansion.",
),
("T7 Style", 5): (
    "Added Canadian spelling enforcement to BASE_PROMPT OUTPUT RULES and extended _replace_em_dash() in agent.py with spelling correction dict.",
    "'behavior' → 'behaviour', 'favorite' → 'favourite', etc. corrected in all responses.",
),
("T7 Style", 6): (
    "Added common misspellings of 'asymptotic' to science keyword list in rag/router.py.",
    "Misspelled variants like 'asymtotic' or 'asymetric' still trigger correct content retrieval.",
),
("T7 Style", 7): (
    "Schedule disclaimer addressed in T2 Change 2B.",
    "Activity schedule details include disclaimer to verify before attending.",
),

}


def main():
    wb = openpyxl.load_workbook(PATH)

    sheet_map = {
        "T1 Intervention": "T1 Intervention",
        "T2 Resource Scope": "T2 Resource Scope",
        "T3 Safety": "T3 Safety",
        "T4 Construct": "T4 Construct",
        "T5 Personalisation": "T5 Personalisation",
        "T6 Source": "T6 Source",
        "T7 Style": "T7 Style",
    }

    wrap = Alignment(wrap_text=True, vertical="top")
    header_font = Font(bold=True)

    for key_sheet, actual_sheet in sheet_map.items():
        ws = wb[actual_sheet]

        # Set headers for new columns I (col 9) and J (col 10)
        for col, label in [(9, "What Was Done"), (10, "New Expected Behaviour")]:
            cell = ws.cell(row=1, column=col)
            cell.value = label
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Set column widths
        ws.column_dimensions["I"].width = 55
        ws.column_dimensions["J"].width = 55

        data_row_index = 1
        for row_num in range(2, ws.max_row + 1):
            if all(ws.cell(row=row_num, column=c).value is None for c in range(1, 7)):
                continue
            lookup = (key_sheet, data_row_index)
            if lookup in DATA:
                done, behaviour = DATA[lookup]
                ci = ws.cell(row=row_num, column=9)
                cj = ws.cell(row=row_num, column=10)
                ci.value = done
                cj.value = behaviour
                ci.alignment = wrap
                cj.alignment = wrap
            else:
                print(f"WARNING: no entry for {lookup}")
            data_row_index += 1

    wb.save(PATH)
    print("Done. Saved to", PATH)


if __name__ == "__main__":
    main()
